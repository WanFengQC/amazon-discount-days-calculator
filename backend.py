from __future__ import annotations

import asyncio
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Literal
import html
import json
import re
import sqlite3
import threading
import time
import urllib.request
import urllib.parse
import urllib.error

from amazon_deal_extractor import (
    DEFAULT_PROFILE_DIR,
    _auto_prepare_offer_context,
    _apply_amazon_us_preferences,
    _build_url,
    _extract_from_html,
    _needs_us_storefront,
    _normalize_us_storefront,
    _read_page_state,
    build_browser_launch_kwargs,
)
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from pydantic import BaseModel, Field
try:
    import chinese_calendar
except Exception:
    chinese_calendar = None


class CalcRequest(BaseModel):
    history_dates: list[str] = Field(default_factory=list)
    future_days: int = Field(ge=1)
    selected_offsets: list[int] = Field(default_factory=list)
    marker_overrides: dict[str, str] = Field(default_factory=dict)
    mode: Literal["even", "centered_even", "centered", "head", "tail", "custom"] = "even"
    preferred_start: int | None = None
    preferred_end: int | None = None
    recommend: bool = False


class DayResult(BaseModel):
    offset: int
    date: str
    selected: bool
    recommended: bool
    available: bool
    violation: bool
    marker: Literal["BD", "PD", "PC"]
    past90_count: int
    remaining_after_select: int


class CalcResponse(BaseModel):
    days: list[DayResult]
    selected_offsets: list[int]
    recommended_offsets: list[int]
    violation_count: int
    message: str


class SharedStateRequest(BaseModel):
    records: list[dict] = Field(default_factory=list)
    product_sim: list[dict] = Field(default_factory=list)
    future_days: int = 90
    mode: str = "even"
    custom_start: int = 0
    custom_end: int = 14
    selected_key: str = ""
    file_name: str = ""
    marker_mode: Literal["BD", "PD", "PC"] = "PD"


class SharedStateResponse(BaseModel):
    records: list[dict] = Field(default_factory=list)
    product_sim: list[dict] = Field(default_factory=list)
    future_days: int = 90
    mode: str = "even"
    custom_start: int = 0
    custom_end: int = 14
    selected_key: str = ""
    file_name: str = ""
    marker_mode: Literal["BD", "PD", "PC"] = "PD"


class FeishuConfigRequest(BaseModel):
    enabled: bool = False
    app_id: str = ""
    app_secret: str = ""
    open_ids: list[str] = Field(default_factory=list)


class FeishuConfigResponse(BaseModel):
    enabled: bool = False
    app_id: str = ""
    open_ids: list[str] = Field(default_factory=list)
    has_secret: bool = False


class ReminderRange(BaseModel):
    start_date: str
    end_date: str
    marker: Literal["BD", "PD", "PC"] = "PC"
    asin: str = ""
    sku: str = ""
    label: str = ""


class ReminderDay(BaseModel):
    remind_date: str
    promo_start: str
    promo_end: str
    items: list[ReminderRange] = Field(default_factory=list)


class AmazonDealRequest(BaseModel):
    url: str
    headers: dict[str, str] = Field(default_factory=dict)
    html: str = ""


class AmazonDealResponse(BaseModel):
    url: str
    discount_type: str | None = None
    discount_strength: str | None = None
    discount_price: str | None = None
    list_price: str | None = None
    typical_price: str | None = None
    regular_price: str | None = None
    prime_member_price: str | None = None


class AsinMonitorConfigRequest(BaseModel):
    enabled: bool = False
    asins: list[str] = Field(default_factory=list)
    us_zip: str = "10001"
    run_hour: int = Field(default=9, ge=0, le=23)
    run_minute: int = Field(default=0, ge=0, le=59)
    headless: bool = True
    retry_count: int = Field(default=2, ge=0, le=5)
    retry_delay_seconds: int = Field(default=20, ge=0, le=300)


class AsinMonitorConfigResponse(BaseModel):
    enabled: bool = False
    asins: list[str] = Field(default_factory=list)
    us_zip: str = "10001"
    run_hour: int = 9
    run_minute: int = 0
    headless: bool = True
    retry_count: int = 2
    retry_delay_seconds: int = 20
    last_run_date: str = ""
    last_run_at: str = ""
    last_auto_run_date: str = ""
    last_auto_run_at: str = ""


class AsinMonitorResult(BaseModel):
    asin: str
    url: str
    fetch_date: str
    fetched_at: str
    attempts: int = 1
    status: str
    reason: str
    crawl_log: str = ""
    html_path: str = ""
    image_url: str = ""
    color: str = ""
    sku: str = ""
    psku: str = ""
    discount_type: str | None = None
    discount_strength: str | None = None
    discount_price: str | None = None
    list_price: str | None = None
    typical_price: str | None = None
    regular_price: str | None = None
    prime_member_price: str | None = None


class AsinMonitorResultsResponse(BaseModel):
    results: list[AsinMonitorResult] = Field(default_factory=list)


class AsinMonitorRunResponse(BaseModel):
    ok: bool = True
    count: int = 0
    results: list[AsinMonitorResult] = Field(default_factory=list)


app = FastAPI(title="Discount Day Calculator API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
BASE_DIR = Path(__file__).resolve().parent
STATE_FILE = BASE_DIR / "shared_state.json"
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
SOURCE_FILE_PATH = DATA_DIR / "source_import.xlsx"
ASIN_SKU_DB_FILE = DATA_DIR / "SKU数据库.xlsx"
LEGACY_ASIN_SKU_DB_FILE = Path.home() / "Downloads" / "SKU数据库.xlsx"
ASIN_MONITOR_HTML_DIR = DATA_DIR / "asin_monitor_html"
ASIN_MONITOR_HTML_DIR.mkdir(parents=True, exist_ok=True)
FEISHU_FILE = DATA_DIR / "feishu_config.json"
REMINDER_LOG_FILE = DATA_DIR / "reminder_log.json"
ASIN_MONITOR_CONFIG_FILE = DATA_DIR / "asin_monitor_config.json"
ASIN_MONITOR_RESULTS_FILE = DATA_DIR / "asin_monitor_results.json"
SQLITE_DB_FILE = DATA_DIR / "app_state.sqlite3"
_reminder_lock = threading.Lock()
_asin_monitor_lock = threading.Lock()
_db_lock = threading.Lock()
_db_initialized = False
_asin_sku_lookup_lock = threading.Lock()
_asin_sku_lookup_cache: dict[str, object] = {
    "path": "",
    "mtime_ns": -1,
    "mapping": {},
}
_asin_monitor_progress: dict = {
    "running": False,
    "total": 0,
    "completed": 0,
    "current_asin": "",
    "status": "idle",
}


@app.get("/", response_class=HTMLResponse)
async def home() -> HTMLResponse:
    html = (BASE_DIR / "index.html").read_text(encoding="utf-8")
    return HTMLResponse(content=html)


def _load_shared_state() -> dict:
    return _db_load_json("shared_state")


def _save_shared_state(data: dict) -> None:
    _db_save_json("shared_state", data)


def _load_json_file(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_json_file(path: Path, data: dict) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _normalize_asin_key(value: object) -> str:
    return str(value or "").strip().upper()


def _clean_cell_text(value: object) -> str:
    return str(value or "").strip()


def _resolve_asin_sku_db_file() -> Path | None:
    for path in (ASIN_SKU_DB_FILE, BASE_DIR / "SKU数据库.xlsx", LEGACY_ASIN_SKU_DB_FILE):
        if path.exists() and path.is_file():
            return path
    return None


def _load_asin_sku_lookup() -> dict[str, dict[str, str]]:
    source = _resolve_asin_sku_db_file()
    if not source:
        return {}
    stat = source.stat()
    cache_key = str(source.resolve())
    with _asin_sku_lookup_lock:
        cached_path = str(_asin_sku_lookup_cache.get("path", "") or "")
        cached_mtime = int(_asin_sku_lookup_cache.get("mtime_ns", -1) or -1)
        if cached_path == cache_key and cached_mtime == int(stat.st_mtime_ns):
            mapping = _asin_sku_lookup_cache.get("mapping", {})
            return mapping if isinstance(mapping, dict) else {}

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if not workbook.sheetnames:
            return {}
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return {}
        headers = {_clean_cell_text(name): idx for idx, name in enumerate(header_row)}
        asin_idx = headers.get("ASIN")
        sku_idx = headers.get("MSKU")
        psku_idx = headers.get("PSKU")
        name_idx = headers.get("品名")
        if asin_idx is None:
            return {}
        mapping: dict[str, dict[str, str]] = {}
        for row in rows:
            if row is None:
                continue
            asin = _normalize_asin_key(row[asin_idx] if asin_idx < len(row) else "")
            if not asin:
                continue
            sku = _clean_cell_text(row[sku_idx] if sku_idx is not None and sku_idx < len(row) else "")
            psku = _clean_cell_text(row[psku_idx] if psku_idx is not None and psku_idx < len(row) else "")
            product_name = _clean_cell_text(row[name_idx] if name_idx is not None and name_idx < len(row) else "")
            mapping[asin] = {
                "sku": sku,
                "psku": psku,
                "product_name": product_name,
            }
    finally:
        workbook.close()

    with _asin_sku_lookup_lock:
        _asin_sku_lookup_cache["path"] = cache_key
        _asin_sku_lookup_cache["mtime_ns"] = int(stat.st_mtime_ns)
        _asin_sku_lookup_cache["mapping"] = mapping
    return mapping


def _enrich_asin_monitor_row(row: dict) -> dict:
    item = dict(row)
    lookup = _load_asin_sku_lookup().get(_normalize_asin_key(item.get("asin")), {})
    item["sku"] = str(lookup.get("sku", "") or "")
    item["psku"] = str(lookup.get("psku", "") or "")
    return item


def _db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(SQLITE_DB_FILE, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def _db_load_json(key: str, default: dict | None = None) -> dict:
    _ensure_sqlite_ready()
    with _db_connect() as conn:
        row = conn.execute("SELECT value FROM kv_store WHERE key = ?", (key,)).fetchone()
    if not row:
        return dict(default or {})
    try:
        value = json.loads(str(row["value"]))
    except Exception:
        return dict(default or {})
    return value if isinstance(value, dict) else dict(default or {})


def _db_save_json(key: str, data: dict) -> None:
    _ensure_sqlite_ready()
    payload = json.dumps(data, ensure_ascii=False)
    with _db_connect() as conn:
        conn.execute(
            """
            INSERT INTO kv_store (key, value, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET
                value = excluded.value,
                updated_at = excluded.updated_at
            """,
            (key, payload, _utc_now_iso()),
        )
        conn.commit()


def _legacy_json_migration_done(conn: sqlite3.Connection) -> bool:
    row = conn.execute("SELECT value FROM meta WHERE key = 'legacy_json_migrated'").fetchone()
    return bool(row and str(row["value"]) == "1")


def _mark_legacy_json_migration_done(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        INSERT INTO meta (key, value, updated_at)
        VALUES ('legacy_json_migrated', '1', ?)
        ON CONFLICT(key) DO UPDATE SET
            value = excluded.value,
            updated_at = excluded.updated_at
        """,
        (_utc_now_iso(),),
    )


def _ensure_asin_monitor_result_columns(conn: sqlite3.Connection) -> None:
    existing = {
        str(row["name"])
        for row in conn.execute("PRAGMA table_info(asin_monitor_results)").fetchall()
    }
    wanted = {
        "list_price": "TEXT",
        "typical_price": "TEXT",
        "regular_price": "TEXT",
        "prime_member_price": "TEXT",
        "reference_price": "TEXT",
        "html_path": "TEXT",
        "image_url": "TEXT",
        "color": "TEXT",
    }
    for name, col_type in wanted.items():
        if name not in existing:
            conn.execute(f"ALTER TABLE asin_monitor_results ADD COLUMN {name} {col_type}")


def _save_asin_monitor_html(asin: str, fetch_date: str, fetched_at: str, html_text: str) -> str:
    asin_key = re.sub(r"[^A-Z0-9_-]+", "_", str(asin or "").strip().upper()) or "UNKNOWN"
    date_key = re.sub(r"[^0-9-]+", "_", str(fetch_date or "").strip()) or date.today().isoformat()
    time_key = re.sub(r"[^0-9T]+", "_", str(fetched_at or "").strip()) or _utc_now_iso().replace(":", "-")
    file_name = f"{asin_key}_{date_key}_{time_key}.html"
    target = ASIN_MONITOR_HTML_DIR / file_name
    target.write_text(html_text or "", encoding="utf-8")
    return str(target.relative_to(BASE_DIR)).replace("\\", "/")


def _migrate_legacy_json_to_sqlite(conn: sqlite3.Connection) -> None:
    legacy_kv = {
        "shared_state": _load_json_file(STATE_FILE),
        "feishu_config": _load_json_file(FEISHU_FILE),
        "reminder_log": _load_json_file(REMINDER_LOG_FILE),
        "asin_monitor_config": _load_json_file(ASIN_MONITOR_CONFIG_FILE),
    }
    for key, value in legacy_kv.items():
        if isinstance(value, dict) and value:
            conn.execute(
                """
                INSERT OR IGNORE INTO kv_store (key, value, updated_at)
                VALUES (?, ?, ?)
                """,
                (key, json.dumps(value, ensure_ascii=False), _utc_now_iso()),
            )

    legacy_results = _load_json_file(ASIN_MONITOR_RESULTS_FILE)
    rows = legacy_results.get("results", []) if isinstance(legacy_results, dict) else []
    for row in rows:
        if not isinstance(row, dict):
            continue
        conn.execute(
            """
            INSERT OR REPLACE INTO asin_monitor_results (
                asin, fetch_date, fetched_at, url, attempts, status, reason,
                crawl_log, html_path, image_url, color, discount_type, discount_strength, discount_price,
                list_price, typical_price, regular_price, prime_member_price, reference_price
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(row.get("asin", "")).strip().upper(),
                str(row.get("fetch_date", "") or ""),
                str(row.get("fetched_at", "") or ""),
                str(row.get("url", "") or ""),
                int(row.get("attempts", 1) or 1),
                str(row.get("status", "") or ""),
                str(row.get("reason", "") or ""),
                str(row.get("crawl_log", "") or ""),
                str(row.get("html_path", "") or ""),
                str(row.get("image_url", "") or ""),
                str(row.get("color", "") or ""),
                row.get("discount_type"),
                row.get("discount_strength"),
                row.get("discount_price"),
                row.get("list_price"),
                row.get("typical_price"),
                row.get("regular_price"),
                row.get("prime_member_price"),
                row.get("reference_price"),
            ),
        )


def _ensure_sqlite_ready() -> None:
    global _db_initialized
    if _db_initialized:
        return
    with _db_lock:
        if _db_initialized:
            return
        with _db_connect() as conn:
            conn.executescript(
                """
                PRAGMA journal_mode=WAL;
                CREATE TABLE IF NOT EXISTS meta (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS kv_store (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS asin_monitor_results (
                    asin TEXT NOT NULL,
                    fetch_date TEXT NOT NULL,
                    fetched_at TEXT NOT NULL,
                    url TEXT NOT NULL,
                    attempts INTEGER NOT NULL DEFAULT 1,
                    status TEXT NOT NULL,
                    reason TEXT NOT NULL,
                    crawl_log TEXT NOT NULL DEFAULT '',
                    html_path TEXT NOT NULL DEFAULT '',
                    image_url TEXT NOT NULL DEFAULT '',
                    color TEXT NOT NULL DEFAULT '',
                    discount_type TEXT,
                    discount_strength TEXT,
                    discount_price TEXT,
                    list_price TEXT,
                    typical_price TEXT,
                    regular_price TEXT,
                    prime_member_price TEXT,
                    reference_price TEXT,
                    PRIMARY KEY (asin, fetch_date)
                );
                CREATE INDEX IF NOT EXISTS idx_asin_monitor_results_fetch_date
                ON asin_monitor_results (fetch_date DESC, asin ASC);
                """
            )
            _ensure_asin_monitor_result_columns(conn)
            if not _legacy_json_migration_done(conn):
                _migrate_legacy_json_to_sqlite(conn)
                _mark_legacy_json_migration_done(conn)
            conn.commit()
        _db_initialized = True


def _utc_now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _default_asin_monitor_config() -> dict:
    return {
        "enabled": False,
        "asins": [],
        "us_zip": "10001",
        "run_hour": 9,
        "run_minute": 0,
        "headless": True,
        "retry_count": 2,
        "retry_delay_seconds": 20,
        "last_run_date": "",
        "last_run_at": "",
        "last_auto_run_date": "",
        "last_auto_run_at": "",
    }


def _normalize_asin_list(items: list[str] | tuple[str, ...] | None) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for item in items or []:
        for raw in re.split(r"[\s,;]+", str(item or "").strip()):
            asin = raw.strip().upper()
            if not asin:
                continue
            if asin in seen:
                continue
            seen.add(asin)
            out.append(asin)
    return out


def _load_asin_monitor_config() -> dict:
    data = _db_load_json("asin_monitor_config")
    base = _default_asin_monitor_config()
    base.update(data if isinstance(data, dict) else {})
    base["asins"] = _normalize_asin_list(base.get("asins", []))
    base["us_zip"] = str(base.get("us_zip", "10001") or "10001").strip() or "10001"
    base["run_hour"] = max(0, min(23, int(base.get("run_hour", 9) or 9)))
    base["run_minute"] = max(0, min(59, int(base.get("run_minute", 0) or 0)))
    base["enabled"] = bool(base.get("enabled", False))
    base["headless"] = bool(base.get("headless", True))
    base["retry_count"] = max(0, min(5, int(base.get("retry_count", 2) or 2)))
    base["retry_delay_seconds"] = max(0, min(300, int(base.get("retry_delay_seconds", 20) or 20)))
    base["last_run_date"] = str(base.get("last_run_date", "") or "")
    base["last_run_at"] = str(base.get("last_run_at", "") or "")
    base["last_auto_run_date"] = str(base.get("last_auto_run_date", "") or "")
    base["last_auto_run_at"] = str(base.get("last_auto_run_at", "") or "")
    return base


def _save_asin_monitor_config(data: dict) -> None:
    payload = _default_asin_monitor_config()
    payload.update(data if isinstance(data, dict) else {})
    payload["asins"] = _normalize_asin_list(payload.get("asins", []))
    _db_save_json("asin_monitor_config", payload)


def _load_asin_monitor_results() -> list[dict]:
    _ensure_sqlite_ready()
    with _db_connect() as conn:
        rows = conn.execute(
            """
            SELECT
                asin,
                url,
                fetch_date,
                fetched_at,
                attempts,
                status,
                reason,
                crawl_log,
                html_path,
                image_url,
                color,
                discount_type,
                discount_strength,
                discount_price,
                list_price,
                typical_price,
                regular_price,
                prime_member_price
            FROM asin_monitor_results
            ORDER BY fetch_date DESC, asin ASC
            """
        ).fetchall()
    out: list[dict] = []
    for row in rows:
        item = dict(row)
        item["html_path"] = str(item.get("html_path", "") or "")
        item["image_url"] = _resolve_asin_monitor_image_url(item)
        item["color"] = _resolve_asin_monitor_color(item)
        out.append(_enrich_asin_monitor_row(item))
    return out


def _save_asin_monitor_results(rows: list[dict]) -> None:
    _ensure_sqlite_ready()
    with _db_connect() as conn:
        conn.execute("DELETE FROM asin_monitor_results")
        for row in rows:
            if not isinstance(row, dict):
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO asin_monitor_results (
                    asin, fetch_date, fetched_at, url, attempts, status, reason,
                    crawl_log, html_path, image_url, color, discount_type, discount_strength, discount_price,
                    list_price, typical_price, regular_price, prime_member_price, reference_price
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(row.get("asin", "")).strip().upper(),
                    str(row.get("fetch_date", "") or ""),
                    str(row.get("fetched_at", "") or ""),
                    str(row.get("url", "") or ""),
                    int(row.get("attempts", 1) or 1),
                    str(row.get("status", "") or ""),
                    str(row.get("reason", "") or ""),
                    str(row.get("crawl_log", "") or ""),
                    str(row.get("html_path", "") or ""),
                    str(row.get("image_url", "") or ""),
                    str(row.get("color", "") or ""),
                    row.get("discount_type"),
                    row.get("discount_strength"),
                    row.get("discount_price"),
                    row.get("list_price"),
                    row.get("typical_price"),
                    row.get("regular_price"),
                    row.get("prime_member_price"),
                    row.get("reference_price"),
                ),
            )
        conn.commit()


def _load_asin_monitor_results_for_date(fetch_date: str) -> list[dict]:
    _ensure_sqlite_ready()
    with _db_connect() as conn:
        rows = conn.execute(
            """
            SELECT
                asin,
                url,
                fetch_date,
                fetched_at,
                attempts,
                status,
                reason,
                crawl_log,
                html_path,
                image_url,
                color,
                discount_type,
                discount_strength,
                discount_price,
                list_price,
                typical_price,
                regular_price,
                prime_member_price
            FROM asin_monitor_results
            WHERE fetch_date = ?
            ORDER BY asin ASC
            """,
            (str(fetch_date),),
        ).fetchall()
    out: list[dict] = []
    for row in rows:
        item = dict(row)
        item["html_path"] = str(item.get("html_path", "") or "")
        item["image_url"] = _resolve_asin_monitor_image_url(item)
        item["color"] = _resolve_asin_monitor_color(item)
        out.append(_enrich_asin_monitor_row(item))
    return out


def _upsert_asin_monitor_results(new_rows: list[dict], keep_latest: int = 1000) -> list[dict]:
    _ensure_sqlite_ready()
    with _db_connect() as conn:
        for row in new_rows:
            if not isinstance(row, dict):
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO asin_monitor_results (
                    asin, fetch_date, fetched_at, url, attempts, status, reason,
                    crawl_log, html_path, image_url, color, discount_type, discount_strength, discount_price,
                    list_price, typical_price, regular_price, prime_member_price, reference_price
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(row.get("asin", "")).strip().upper(),
                    str(row.get("fetch_date", "") or ""),
                    str(row.get("fetched_at", "") or ""),
                    str(row.get("url", "") or ""),
                    int(row.get("attempts", 1) or 1),
                    str(row.get("status", "") or ""),
                    str(row.get("reason", "") or ""),
                    str(row.get("crawl_log", "") or ""),
                    str(row.get("html_path", "") or ""),
                    str(row.get("image_url", "") or ""),
                    str(row.get("color", "") or ""),
                    row.get("discount_type"),
                    row.get("discount_strength"),
                    row.get("discount_price"),
                    row.get("list_price"),
                    row.get("typical_price"),
                    row.get("regular_price"),
                    row.get("prime_member_price"),
                    row.get("reference_price"),
                ),
            )
        stale = conn.execute(
            """
            SELECT asin, fetch_date
            FROM asin_monitor_results
            ORDER BY fetch_date DESC, asin ASC
            LIMIT -1 OFFSET ?
            """,
            (max(0, keep_latest),),
        ).fetchall()
        for stale_row in stale:
            conn.execute(
                "DELETE FROM asin_monitor_results WHERE asin = ? AND fetch_date = ?",
                (str(stale_row["asin"]), str(stale_row["fetch_date"])),
            )
        conn.commit()
    return _load_asin_monitor_results()


def _parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def _extract_first_group(text: str, patterns: list[str], flags: int = re.IGNORECASE | re.DOTALL) -> str | None:
    for pattern in patterns:
        m = re.search(pattern, text, flags)
        if m:
            value = m.group(1)
            value = html.unescape(value)
            value = re.sub(r"\s+", " ", value).strip()
            if value:
                return value
    return None


def _normalize_price(value: str | None) -> str | None:
    if not value:
        return None
    text = value.replace(",", "").strip()
    m = re.search(r"(\$?\d+(?:\.\d{1,2})?)", text)
    if not m:
        return None
    amount = m.group(1)
    if not amount.startswith("$"):
        amount = f"${amount}"
    return amount


def _price_to_number(value: str | None) -> float | None:
    if not value:
        return None
    match = re.search(r"(\d[\d,]*(?:\.\d{2})?)", value)
    if not match:
        return None
    try:
        return float(match.group(1).replace(",", ""))
    except ValueError:
        return None


def _infer_discount_strength(
    discount_price: str | None,
    list_price: str | None,
    typical_price: str | None,
    regular_price: str | None,
) -> str | None:
    sale = _price_to_number(discount_price)
    if not sale or sale <= 0:
        return None
    for candidate in (list_price, regular_price, typical_price):
        reference = _price_to_number(candidate)
        if not reference or reference <= sale:
            continue
        pct = round((reference - sale) / reference * 100)
        if pct > 0:
            return f"-{pct}%"
    return None


def _clean_discount_strength(value: str | None) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if re.search(r"\b(off|qualify|coupon)\b", text, re.IGNORECASE):
        return None
    match = re.fullmatch(r"(-?\d+%)", text)
    if not match:
        return None
    normalized = match.group(1)
    return normalized if normalized.startswith("-") else f"-{normalized}"


def _has_public_discount_signal(
    discount_type: str | None,
    discount_strength: str | None,
    discount_price: str | None,
    list_price: str | None,
    typical_price: str | None,
    regular_price: str | None,
) -> bool:
    if str(discount_type or "").strip():
        return True
    if str(discount_strength or "").strip():
        return True
    sale = _price_to_number(discount_price)
    if not sale:
        return False
    for candidate in (list_price, typical_price, regular_price):
        reference = _price_to_number(candidate)
        if reference and reference > sale:
            return True
    return False


def _finalize_price_fields(
    discount_type: str | None,
    discount_strength: str | None,
    discount_price: str | None,
    list_price: str | None,
    typical_price: str | None,
    regular_price: str | None,
) -> tuple[str | None, str | None, str | None]:
    sale = _price_to_number(discount_price)
    regular = _price_to_number(regular_price)
    if sale and regular and sale == regular:
        discount_price = None
        discount_strength = None
        discount_type = None
    if discount_price and not _has_public_discount_signal(
        discount_type,
        discount_strength,
        discount_price,
        list_price,
        typical_price,
        regular_price,
    ):
        if not regular_price:
            regular_price = discount_price
        discount_price = None
        discount_strength = None
    return discount_strength, discount_price, regular_price


def _resolve_asin_monitor_image_url(row: dict) -> str:
    image_url = str(row.get("image_url", "") or "").strip()
    if image_url:
        return image_url
    html_path = str(row.get("html_path", "") or "").strip()
    if not html_path:
        return ""
    target = (BASE_DIR / html_path).resolve()
    try:
        target.relative_to(BASE_DIR.resolve())
    except ValueError:
        return ""
    if not target.exists() or not target.is_file():
        return ""
    try:
        html_text = target.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return ""
    parsed = _extract_from_html(html_text, "")
    return str(parsed.get("image_url", "") or "").strip()


def _resolve_asin_monitor_color(row: dict) -> str:
    color = str(row.get("color", "") or "").strip()
    if color:
        return color
    html_path = str(row.get("html_path", "") or "").strip()
    if not html_path:
        return ""
    target = (BASE_DIR / html_path).resolve()
    try:
        target.relative_to(BASE_DIR.resolve())
    except ValueError:
        return ""
    if not target.exists() or not target.is_file():
        return ""
    try:
        html_text = target.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return ""
    parsed = _extract_from_html(html_text, "")
    return str(parsed.get("color", "") or "").strip()


def _fetch_amazon_page(url: str, extra_headers: dict[str, str] | None = None) -> str:
    headers = {
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/148.0.0.0 Safari/537.36 Edg/148.0.0.0"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    }
    if extra_headers:
        headers.update({str(k): str(v) for k, v in extra_headers.items() if str(k).strip()})
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=25) as resp:
        charset = resp.headers.get_content_charset() or "utf-8"
        body = resp.read()
    return body.decode(charset, errors="replace")


def _parse_amazon_deal_fields(page_html: str) -> dict:
    page_text = html.unescape(re.sub(r"<[^>]+>", " ", page_html))
    page_text = re.sub(r"\s+", " ", page_text).strip()
    extracted = _extract_from_html(page_html, page_text)
    return {
        "discount_type": extracted.get("discount_type"),
        "discount_strength": extracted.get("discount_strength"),
        "discount_price": extracted.get("discount_price"),
        "list_price": extracted.get("list_price"),
        "typical_price": extracted.get("typical_price"),
        "regular_price": extracted.get("regular_price"),
        "prime_member_price": extracted.get("prime_member_price"),
    }


def _run_asin_monitor_once(config: dict) -> list[dict]:
    asins = _normalize_asin_list(config.get("asins", []))
    if not asins:
        return []
    profile_name = "headless" if bool(config.get("headless", True)) else "headed"
    profile_dir = Path(DEFAULT_PROFILE_DIR).resolve() / profile_name
    profile_dir.mkdir(parents=True, exist_ok=True)
    us_zip = str(config.get("us_zip", "10001") or "10001").strip() or "10001"
    headless = bool(config.get("headless", True))
    retry_count = max(0, int(config.get("retry_count", 2) or 0))
    retry_delay_seconds = max(0, int(config.get("retry_delay_seconds", 20) or 0))
    today_key = date.today().isoformat()
    now_key = _utc_now_iso()
    rows: list[dict] = []

    with sync_playwright() as p:
        try:
            launch_kwargs = dict(
                user_data_dir=str(profile_dir),
                headless=headless,
                locale="en-US",
                timezone_id="America/Los_Angeles",
                viewport={"width": 1600, "height": 1400},
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/148.0.0.0 Safari/537.36"
                ),
            )
            try:
                context = p.chromium.launch_persistent_context(
                    **build_browser_launch_kwargs(),
                    **launch_kwargs,
                )
            except Exception as e:
                if "Chromium distribution 'msedge' is not found" not in str(e):
                    raise
                context = p.chromium.launch_persistent_context(**launch_kwargs)
        except Exception as e:
            raise RuntimeError(
                f"failed to open persistent browser profile ({profile_name}); "
                f"close existing browser windows using this monitor profile and retry: {e}"
            ) from e
        try:
            _apply_amazon_us_preferences(context)
            page = context.pages[0] if context.pages else context.new_page()
            total = len(asins)
            for idx, asin in enumerate(asins):
                _asin_monitor_progress.update({
                    "running": True,
                    "total": total,
                    "completed": idx,
                    "current_asin": asin,
                    "status": f"正在抓取 {asin} ({idx + 1}/{total})",
                })
                url = _build_url(asin, None)
                attempts = 0
                row: dict | None = None
                attempt_logs: list[str] = []
                while attempts <= retry_count:
                    attempts += 1
                    try:
                        attempt_logs.append(f"[attempt {attempts}] open {url}")
                        page.goto(url, wait_until="domcontentloaded", timeout=90000)
                        page.wait_for_timeout(5000)
                        attempt_logs.append(f"[attempt {attempts}] page loaded: {page.url}")
                        _auto_prepare_offer_context(page, us_zip)
                        page.wait_for_timeout(3000)
                        precheck_state = _read_page_state(page)
                        if _needs_us_storefront(precheck_state):
                            attempt_logs.append(
                                f"[attempt {attempts}] detected localized storefront or unavailable localized offer; forcing en-US/USD storefront"
                            )
                            _normalize_us_storefront(page, url, us_zip)
                        state = _read_page_state(page)
                        parsed = _extract_from_html(state.html, state.body_text)
                        fetched_at = _utc_now_iso()
                        html_path = _save_asin_monitor_html(asin, today_key, fetched_at, state.html)
                        attempt_logs.append(
                            f"[attempt {attempts}] status={state.status}; reason={state.reason}; "
                            f"image_url={parsed.get('image_url') or '-'}; "
                            f"color={parsed.get('color') or '-'}; "
                            f"discount_type={parsed.get('discount_type') or '-'}; "
                            f"discount_strength={parsed.get('discount_strength') or '-'}; "
                            f"discount_price={parsed.get('discount_price') or '-'}; "
                            f"list_price={parsed.get('list_price') or '-'}; "
                            f"typical_price={parsed.get('typical_price') or '-'}; "
                            f"regular_price={parsed.get('regular_price') or '-'}; "
                            f"prime_member_price={parsed.get('prime_member_price') or '-'}"
                        )
                        row = {
                            "asin": asin,
                            "url": page.url,
                            "fetch_date": today_key,
                            "fetched_at": fetched_at,
                            "attempts": attempts,
                            "status": state.status,
                            "reason": state.reason,
                            "crawl_log": "\n".join(attempt_logs),
                            "html_path": html_path,
                            "image_url": parsed.get("image_url") or "",
                            "color": parsed.get("color") or "",
                            "discount_type": parsed.get("discount_type"),
                            "discount_strength": parsed.get("discount_strength"),
                            "discount_price": parsed.get("discount_price"),
                            "list_price": parsed.get("list_price"),
                            "typical_price": parsed.get("typical_price"),
                            "regular_price": parsed.get("regular_price"),
                            "prime_member_price": parsed.get("prime_member_price"),
                            "reference_price": None,
                        }
                        should_retry = state.status in {"blocked", "unknown"} and attempts <= retry_count
                        if should_retry and retry_delay_seconds > 0:
                            attempt_logs.append(
                                f"[attempt {attempts}] retry scheduled after {retry_delay_seconds}s"
                            )
                            time.sleep(retry_delay_seconds)
                            continue
                        break
                    except Exception as e:
                        attempt_logs.append(f"[attempt {attempts}] exception: {e}")
                        fetched_at = _utc_now_iso()
                        html_path = ""
                        try:
                            html_snapshot = page.content()
                            if html_snapshot:
                                html_path = _save_asin_monitor_html(asin, today_key, fetched_at, html_snapshot)
                        except Exception:
                            html_path = ""
                        row = {
                            "asin": asin,
                            "url": url,
                            "fetch_date": today_key,
                            "fetched_at": fetched_at,
                            "attempts": attempts,
                            "status": "error",
                            "reason": str(e),
                            "crawl_log": "\n".join(attempt_logs),
                            "html_path": html_path,
                            "image_url": "",
                            "color": "",
                            "discount_type": None,
                            "discount_strength": None,
                            "discount_price": None,
                            "list_price": None,
                            "typical_price": None,
                            "regular_price": None,
                            "prime_member_price": None,
                            "reference_price": None,
                        }
                        if attempts <= retry_count and retry_delay_seconds > 0:
                            attempt_logs.append(
                                f"[attempt {attempts}] retry scheduled after {retry_delay_seconds}s"
                            )
                            time.sleep(retry_delay_seconds)
                            continue
                        break
                if row is None:
                    row = {
                        "asin": asin,
                        "url": url,
                        "fetch_date": today_key,
                        "fetched_at": now_key,
                        "attempts": attempts or 1,
                        "status": "error",
                        "reason": "unknown crawler failure",
                        "crawl_log": "\n".join(attempt_logs),
                        "html_path": "",
                        "image_url": "",
                        "color": "",
                        "discount_type": None,
                        "discount_strength": None,
                        "discount_price": None,
                        "list_price": None,
                        "typical_price": None,
                        "regular_price": None,
                        "prime_member_price": None,
                        "reference_price": None,
                    }
                rows.append(row)
                _asin_monitor_progress.update({
                    "running": True,
                    "total": total,
                    "completed": idx + 1,
                    "current_asin": asin,
                    "status": f"已完成 {asin} ({idx + 1}/{total})",
                })
        finally:
            context.close()
    _asin_monitor_progress.update({
        "running": False,
        "total": total,
        "completed": total,
        "current_asin": "",
        "status": "done",
    })
    return rows


def _execute_asin_monitor_run(force: bool = False, trigger: str = "manual") -> list[dict]:
    with _asin_monitor_lock:
        config = _load_asin_monitor_config()
        today_key = date.today().isoformat()
        configured_asins = _normalize_asin_list(config.get("asins", []))
        existing_rows = _load_asin_monitor_results_for_date(today_key)
        existing_asins = {
            str(row.get("asin", "")).strip().upper()
            for row in existing_rows
            if str(row.get("asin", "")).strip()
        }

        if trigger == "auto":
            target_asins = configured_asins if force else [asin for asin in configured_asins if asin not in existing_asins]
        else:
            dedupe_key = "last_run_date"
            if not force and str(config.get(dedupe_key, "")) == today_key and set(configured_asins).issubset(existing_asins):
                return existing_rows
            target_asins = configured_asins

        if not target_asins:
            if trigger == "auto":
                config["last_auto_run_date"] = today_key
                config["last_auto_run_at"] = _utc_now_iso()
                _save_asin_monitor_config(config)
            return existing_rows

        run_config = dict(config)
        run_config["asins"] = target_asins
        rows = _run_asin_monitor_once(run_config)
        _upsert_asin_monitor_results(rows)
        config["last_run_date"] = today_key
        config["last_run_at"] = _utc_now_iso()
        if trigger == "auto":
            config["last_auto_run_date"] = today_key
            config["last_auto_run_at"] = config["last_run_at"]
        _save_asin_monitor_config(config)
        return [_enrich_asin_monitor_row(row) for row in rows]


def _asin_monitor_worker() -> None:
    while True:
        try:
            config = _load_asin_monitor_config()
            if config.get("enabled") and config.get("asins"):
                now = datetime.now()
                run_after = now.replace(
                    hour=int(config.get("run_hour", 9) or 9),
                    minute=int(config.get("run_minute", 0) or 0),
                    second=0,
                    microsecond=0,
                )
                if now >= run_after:
                    _execute_asin_monitor_run(force=False, trigger="auto")
        except Exception:
            pass
        time.sleep(30)


def _build_dates(today: date, n: int) -> list[date]:
    return [today + timedelta(days=i) for i in range(n)]


def _build_marker_list(
    future_days: int,
    selected: list[bool],
    mode: str,
    preferred_start: int | None,
    preferred_end: int | None,
) -> list[str]:
    markers = ["PC"] * future_days
    s = 0 if preferred_start is None else max(0, min(future_days - 1, preferred_start))
    e = future_days - 1 if preferred_end is None else max(0, min(future_days - 1, preferred_end))
    l, r = (s, e) if s <= e else (e, s)
    for i in range(future_days):
        if not selected[i]:
            markers[i] = "PC"
            continue
        if mode == "custom" and l <= i <= r:
            markers[i] = "BD"
        else:
            markers[i] = "PD"
    return markers


def _normalize_marker(v: str | None) -> str | None:
    if not v:
        return None
    x = str(v).strip().upper()
    if x in {"BD", "PD", "PC"}:
        return x
    return None


def _compute(
    history_dates: list[str],
    future_days: int,
    selected_offsets: list[int],
    marker_overrides: dict[str, str],
    mode: str,
    preferred_start: int | None,
    preferred_end: int | None,
    recommend: bool,
) -> CalcResponse:
    today = date.today()
    future_dates = _build_dates(today, future_days)
    history_set = {_parse_date(x) for x in history_dates}

    # history contribution for each future day, in window [day-89, day], history only counts before today
    hist_count = [0] * future_days
    for i, d in enumerate(future_dates):
        c = 0
        for k in range(90):
            t = d - timedelta(days=k)
            if t < today and t in history_set:
                c += 1
        hist_count[i] = c

    selected = [False] * future_days
    for x in selected_offsets:
        if 0 <= x < future_days:
            selected[x] = True

    def selected_window_counts(sel: list[bool]) -> list[int]:
        pref = [0] * (future_days + 1)
        for i in range(future_days):
            pref[i + 1] = pref[i] + (1 if sel[i] else 0)
        out = [0] * future_days
        for i in range(future_days):
            l = 0 if i < 89 else i - 89
            out[i] = pref[i + 1] - pref[l]
        return out

    def validate(sel: list[bool]) -> tuple[bool, list[bool]]:
        sw = selected_window_counts(sel)
        viol = [False] * future_days
        for i in range(future_days):
            if sel[i] and (hist_count[i] + sw[i] >= 45):
                viol[i] = True
        return (not any(viol), viol)

    def can_select(idx: int, sel: list[bool]) -> bool:
        if sel[idx]:
            return True
        trial = sel[:]
        trial[idx] = True
        ok, _ = validate(trial)
        return ok

    def apply_block(start: int, seed: list[bool]) -> list[bool]:
        out = seed[:]
        for i in range(start, start + 7):
            out[i] = True
        return out

    def can_select_block(start: int, seed: list[bool], chosen_starts: list[int] | None = None) -> bool:
        if future_days < 7 or start < 0 or start + 6 >= future_days:
            return False
        if chosen_starts:
            for s in chosen_starts:
                if abs(start - s) < 7:
                    return False
        trial = apply_block(start, seed)
        ok, _ = validate(trial)
        return ok

    def recommend_even(sel_seed: list[bool]) -> list[bool]:
        # 严格7天连续 + 全区间分散：每个促销段必须连续7天，段与段不重叠
        sel = sel_seed[:]
        n = future_days
        if n < 7:
            return sel

        block_starts = list(range(0, n - 6))

        def can_select_block(start: int, seed: list[bool], chosen_starts: list[int]) -> bool:
            # 防止7天段重叠，避免“挤在一堆”
            for s in chosen_starts:
                if abs(start - s) < 7:
                    return False
            trial = seed[:]
            for i in range(start, start + 7):
                trial[i] = True
            ok, _ = validate(trial)
            return ok

        def apply_block(start: int, seed: list[bool]) -> list[bool]:
            out = seed[:]
            for i in range(start, start + 7):
                out[i] = True
            return out

        # 先估计最多可放几个7天段
        max_blocks = 0
        probe = sel[:]
        probe_starts: list[int] = []
        for s in block_starts:
            if can_select_block(s, probe, probe_starts):
                probe = apply_block(s, probe)
                probe_starts.append(s)
                max_blocks += 1

        # 从最多段数向下找首个可行的“全区间均匀”布局
        for k in range(max_blocks, 0, -1):
            trial = sel[:]
            chosen_starts: list[int] = []
            ok_all = True

            # 以未来总天数做等距锚点（按段中心点分布）
            anchors = [round(((j + 0.5) * n) / k - 0.5) for j in range(k)]
            for a in anchors:
                # 把锚点映射到可能的块起点
                near = []
                for s in block_starts:
                    center = s + 3
                    near.append((abs(center - a), s))
                near.sort(key=lambda x: x[0])

                picked = -1
                for _, s in near:
                    if can_select_block(s, trial, chosen_starts):
                        picked = s
                        break
                if picked == -1:
                    ok_all = False
                    break
                chosen_starts.append(picked)
                trial = apply_block(picked, trial)

            if ok_all:
                return trial

        return sel

    def recommend_centered_even(sel_seed: list[bool]) -> list[bool]:
        # 中心均匀也遵守7天连续块
        sel = sel_seed[:]
        n = future_days
        if n < 7:
            return sel
        starts = list(range(0, n - 6))
        center = (n - 1) / 2
        chosen_starts: list[int] = []
        while True:
            best = -1
            best_score = -10**9
            for s in starts:
                if not can_select_block(s, sel, chosen_starts):
                    continue
                block_center = s + 3
                # 兼顾分散和中心
                if chosen_starts:
                    min_dist = min(abs(s - x) for x in chosen_starts)
                else:
                    min_dist = n
                score = min_dist * 100 - abs(block_center - center)
                if score > best_score:
                    best_score = score
                    best = s
            if best == -1:
                break
            chosen_starts.append(best)
            sel = apply_block(best, sel)
        return sel

    def recommend_centered(sel_seed: list[bool]) -> list[bool]:
        sel = sel_seed[:]
        n = future_days
        center = (n - 1) / 2
        order = sorted(range(n), key=lambda i: (abs(i - center), i))
        for i in order:
            if sel[i]:
                continue
            if can_select(i, sel):
                sel[i] = True
        return sel

    def recommend_ordered(sel_seed: list[bool], order: list[int]) -> list[bool]:
        sel = sel_seed[:]
        for i in order:
            if i < 0 or i >= future_days:
                continue
            if sel[i]:
                continue
            if can_select(i, sel):
                sel[i] = True
        return sel

    def recommend_even_with_priority(sel_seed: list[bool], preferred: set[int]) -> list[bool]:
        # 自定义区间均匀分布：也按7天连续块
        sel = sel_seed[:]
        n = future_days
        if n < 7:
            return sel

        all_starts = list(range(0, n - 6))
        pref_starts = [s for s in all_starts if all((s + k) in preferred for k in range(7))]
        chosen_starts: list[int] = []

        # Phase 1: 自定义区间内尽量全选（按区间内均匀）
        while True:
            best = -1
            best_score = -10**9
            if pref_starts:
                center = (pref_starts[0] + pref_starts[-1]) / 2
            else:
                center = 0
            for s in pref_starts:
                if not can_select_block(s, sel, chosen_starts):
                    continue
                if chosen_starts:
                    min_dist = min(abs(s - x) for x in chosen_starts)
                else:
                    min_dist = n
                score = min_dist * 100 - abs(s - center)
                if score > best_score:
                    best_score = score
                    best = s
            if best == -1:
                break
            chosen_starts.append(best)
            sel = apply_block(best, sel)

        # Phase 2: 剩余块在全局均匀分布
        while True:
            best = -1
            best_score = -10**9
            global_center = (n - 1) / 2
            for s in all_starts:
                if not can_select_block(s, sel, chosen_starts):
                    continue
                block_center = s + 3
                if chosen_starts:
                    min_dist = min(abs(s - x) for x in chosen_starts)
                else:
                    min_dist = n
                score = min_dist * 100 - abs(block_center - global_center)
                if score > best_score:
                    best_score = score
                    best = s
            if best == -1:
                break
            chosen_starts.append(best)
            sel = apply_block(best, sel)
        return sel

    recommended_mask = [False] * future_days
    if recommend:
        if mode == "centered_even":
            selected = recommend_centered_even(selected)
        elif mode == "centered":
            selected = recommend_centered(selected)
        elif mode == "head":
            selected = recommend_ordered(selected, list(range(future_days)))
        elif mode == "tail":
            selected = recommend_ordered(selected, list(range(future_days - 1, -1, -1)))
        elif mode == "custom":
            s = 0 if preferred_start is None else max(0, min(future_days - 1, preferred_start))
            e = future_days - 1 if preferred_end is None else max(0, min(future_days - 1, preferred_end))
            if s <= e:
                pref = set(range(s, e + 1))
            else:
                pref = set(range(e, s + 1))
            selected = recommend_even_with_priority(selected, pref)
        else:
            selected = recommend_even(selected)
        recommended_mask = selected[:]

    ok, violations = validate(selected)
    sw = selected_window_counts(selected)
    markers = _build_marker_list(future_days, selected, mode, preferred_start, preferred_end)
    for k, v in (marker_overrides or {}).items():
        try:
            idx = int(k)
        except Exception:
            continue
        if idx < 0 or idx >= future_days or not selected[idx]:
            continue
        mk = _normalize_marker(v)
        if mk in {"BD", "PD"}:
            markers[idx] = mk

    availability = [False] * future_days
    for i in range(future_days):
        availability[i] = selected[i] or can_select(i, selected)

    days: list[DayResult] = []
    for i, d in enumerate(future_dates):
        used = hist_count[i] + sw[i]
        days.append(
            DayResult(
                offset=i,
                date=d.isoformat(),
                selected=selected[i],
                recommended=recommended_mask[i],
                available=availability[i],
                violation=violations[i],
                marker=markers[i],
                past90_count=used,
                remaining_after_select=max(0, 44 - used),
            )
        )

    violation_count = sum(1 for x in violations if x)
    if ok:
        msg = "当前方案满足 90 天窗口约束。"
    else:
        msg = f"检测到 {violation_count} 个已选日期不满足“该日前90天<45天”，请调整。"

    return CalcResponse(
        days=days,
        selected_offsets=[i for i, v in enumerate(selected) if v],
        recommended_offsets=[i for i, v in enumerate(recommended_mask) if v],
        violation_count=violation_count,
        message=msg,
    )


def _group_contiguous(offsets: list[int]) -> list[tuple[int, int]]:
    if not offsets:
        return []
    arr = sorted(set(int(x) for x in offsets if isinstance(x, int) or (isinstance(x, str) and x.isdigit())))
    if not arr:
        return []
    out: list[tuple[int, int]] = []
    s = arr[0]
    e = arr[0]
    for x in arr[1:]:
        if x == e + 1:
            e = x
        else:
            out.append((s, e))
            s = x
            e = x
    out.append((s, e))
    return out


def _shift_weekend_to_friday(d: date) -> date:
    w = d.weekday()
    if w == 5:
        return d - timedelta(days=1)
    if w == 6:
        return d - timedelta(days=2)
    return d


def _is_cn_holiday(d: date) -> bool:
    if chinese_calendar is None:
        return False
    try:
        return bool(chinese_calendar.is_holiday(d))
    except Exception:
        return False


def _shift_to_previous_workday(d: date) -> date:
    # 规则：前两天若遇周末或法定节假日，持续前移到最近工作日
    x = d
    while True:
        if x.weekday() >= 5:
            x -= timedelta(days=1)
            continue
        if _is_cn_holiday(x):
            x -= timedelta(days=1)
            continue
        break
    return x


def _build_reminder_days(shared: dict, today: date | None = None) -> list[dict]:
    if today is None:
        today = date.today()
    records = {str(x.get("key", "")): x for x in shared.get("records", []) if isinstance(x, dict)}
    future_days = max(1, int(shared.get("future_days", 90) or 90))
    mode = str(shared.get("mode", "even") or "even")
    preferred_start = shared.get("custom_start", 0)
    preferred_end = shared.get("custom_end", 14)
    grouped: dict[str, dict] = {}
    for sim in shared.get("product_sim", []):
        if not isinstance(sim, dict):
            continue
        key = str(sim.get("key", "") or "")
        if not key:
            continue
        raw_offsets = sim.get("selectedOffsets", []) or []
        selected_markers = sim.get("selectedMarkers", {}) or {}
        selected = [False] * future_days
        for x in raw_offsets:
            try:
                ix = int(x)
            except Exception:
                continue
            if 0 <= ix < future_days:
                selected[ix] = True
        markers = _build_marker_list(future_days, selected, mode, preferred_start, preferred_end)
        for k, v in selected_markers.items():
            try:
                idx = int(k)
            except Exception:
                continue
            if idx < 0 or idx >= future_days or not selected[idx]:
                continue
            mk = _normalize_marker(v)
            if mk in {"BD", "PD"}:
                markers[idx] = mk
        marker_ranges: list[tuple[int, int, str]] = []
        start = 0
        cur = markers[0]
        for i in range(1, future_days):
            if markers[i] != cur:
                marker_ranges.append((start, i - 1, cur))
                start = i
                cur = markers[i]
        marker_ranges.append((start, future_days - 1, cur))
        rec = records.get(key, {})
        meta = rec.get("meta", {}) if isinstance(rec.get("meta"), dict) else {}
        label = str(rec.get("label", key) or key)
        asin = str(meta.get("ASIN", "") or "")
        sku = str(meta.get("SKU", "") or "")
        for s, e, marker in marker_ranges:
            if s < 0:
                continue
            mark_start = today + timedelta(days=s)
            mark_end = today + timedelta(days=e)
            remind_date = _shift_to_previous_workday(mark_start - timedelta(days=2))
            k = remind_date.isoformat()
            if k not in grouped:
                grouped[k] = {
                    "remind_date": k,
                    "promo_start": mark_start.isoformat(),
                    "promo_end": mark_end.isoformat(),
                    "items": [],
                }
            grouped[k]["items"].append(
                {
                    "start_date": mark_start.isoformat(),
                    "end_date": mark_end.isoformat(),
                    "marker": marker,
                    "asin": asin,
                    "sku": sku,
                    "label": label,
                }
            )
    def _sort_key(x: dict) -> tuple[str, str]:
        return (x["remind_date"], x["promo_start"])
    return sorted(grouped.values(), key=_sort_key)


def _feishu_get_token(app_id: str, app_secret: str) -> str:
    body = json.dumps({"app_id": app_id, "app_secret": app_secret}, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        data=body,
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        raw = resp.read().decode("utf-8")
    data = json.loads(raw)
    if int(data.get("code", -1)) != 0:
        raise RuntimeError(f"飞书取 token 失败: {data.get('msg', 'unknown error')}")
    token = str(data.get("tenant_access_token", "") or "")
    if not token:
        raise RuntimeError("飞书 tenant_access_token 为空")
    return token


def _send_feishu_message(token: str, open_id: str, msg_type: str, content_obj: dict) -> None:
    payload = json.dumps(
        {
            "receive_id": open_id,
            "msg_type": msg_type,
            "content": json.dumps(content_obj, ensure_ascii=False),
        },
        ensure_ascii=False,
    ).encode("utf-8")
    req = urllib.request.Request(
        "https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=open_id",
        data=payload,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "Authorization": f"Bearer {token}",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        raw = resp.read().decode("utf-8")
    data = json.loads(raw)
    if int(data.get("code", -1)) != 0:
        raise RuntimeError(f"飞书发消息失败: {data.get('msg', 'unknown error')}")


def _feishu_get_json(token: str, url: str) -> dict:
    req = urllib.request.Request(
        url,
        headers={"Authorization": f"Bearer {token}"},
        method="GET",
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            raw = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        raw = ""
        try:
            raw = e.read().decode("utf-8")
        except Exception:
            raw = ""
        if raw:
            try:
                err = json.loads(raw)
                raise RuntimeError(f"飞书接口失败 code={err.get('code')} msg={err.get('msg')} detail={err}")
            except Exception:
                raise RuntimeError(f"飞书接口失败 HTTP {e.code}: {raw}")
        raise RuntimeError(f"飞书接口失败 HTTP {e.code}")
    data = json.loads(raw)
    if int(data.get("code", -1)) != 0:
        raise RuntimeError(f"飞书接口失败 code={data.get('code')} msg={data.get('msg')} detail={data}")
    return data.get("data", {}) or {}


def _feishu_get_user_profile(token: str, open_id: str) -> dict:
    url = f"https://open.feishu.cn/open-apis/contact/v3/users/{urllib.parse.quote(open_id)}?user_id_type=open_id"
    data = _feishu_get_json(token, url)
    user = data.get("user", {}) if isinstance(data, dict) else {}
    return {
        "open_id": open_id,
        "name": str(user.get("name", "") or user.get("en_name", "") or user.get("nickname", "") or f"成员-{str(user.get('user_id', ''))[-4:]}"),
        "avatar": str(user.get("avatar", {}).get("avatar_72", "") if isinstance(user.get("avatar"), dict) else ""),
        "user_id": str(user.get("user_id", "") or ""),
    }


def _feishu_get_users_batch(token: str, open_ids: list[str]) -> list[dict]:
    params: list[tuple[str, str]] = [
        ("user_id_type", "open_id"),
        ("department_id_type", "open_department_id"),
    ]
    for oid in open_ids:
        params.append(("user_ids", oid))
    url = "https://open.feishu.cn/open-apis/contact/v3/users/batch?" + urllib.parse.urlencode(params)
    data = _feishu_get_json(token, url)
    items = data.get("items", []) if isinstance(data, dict) else []
    out: list[dict] = []
    for it in items:
        avatar_obj = it.get("avatar", {}) if isinstance(it.get("avatar"), dict) else {}
        avatar = str(
            avatar_obj.get("avatar_72")
            or avatar_obj.get("avatar_240")
            or avatar_obj.get("avatar_origin")
            or it.get("avatar_url")
            or ""
        )
        out.append(
            {
                "open_id": str(it.get("open_id", "") or ""),
                "name": str(
                    it.get("name", "")
                    or it.get("nickname", "")
                    or it.get("en_name", "")
                    or f"成员-{str(it.get('user_id', ''))[-4:]}"
                ),
                "avatar": avatar,
                "user_id": str(it.get("user_id", "") or ""),
            }
        )
    return out


def _build_message(day_row: dict) -> str:
    lines = [
        f"?????{day_row['promo_start']} - {day_row['promo_end']} ??????",
        f"?????{day_row['remind_date']}",
        "?????",
    ]
    seen = set()
    for item in day_row.get("items", []):
        asin = str(item.get("asin", "") or "").strip()
        sku = str(item.get("sku", "") or "").strip()
        label = str(item.get("label", "") or "").strip()
        start = str(item.get("start_date", "") or "")
        end = str(item.get("end_date", "") or "")
        marker = str(item.get("marker", "PC") or "PC").upper()
        base = label if label else f"{asin} | {sku}".strip(" |")
        base = base if base else "-"
        key = (base, start, end, marker)
        if key in seen:
            continue
        seen.add(key)
        lines.append(f"- [{marker}] {base} | {start}~{end}")
    return "\n".join(lines)


def _build_card_content(day_row: dict) -> dict:
    seen = set()
    item_lines: list[str] = []
    for item in day_row.get("items", []):
        asin = str(item.get("asin", "") or "").strip()
        sku = str(item.get("sku", "") or "").strip()
        label = str(item.get("label", "") or "").strip()
        start = str(item.get("start_date", "") or "")
        end = str(item.get("end_date", "") or "")
        marker = str(item.get("marker", "PC") or "PC").upper()
        base = label if label else f"{asin} | {sku}".strip(" |")
        base = base if base else "-"
        key = (base, start, end, marker)
        if key in seen:
            continue
        seen.add(key)
        item_lines.append(f"[{marker}] {base} | {start}~{end}")
    text = "\n".join(item_lines) if item_lines else "-"
    return {
        "config": {"wide_screen_mode": True},
        "header": {
            "template": "blue",
            "title": {"tag": "plain_text", "content": f"???{day_row.get('promo_start')} ~ {day_row.get('promo_end')}"},
        },
        "elements": [
            {"tag": "markdown", "content": f"**????**?{day_row.get('remind_date')}"},
            {"tag": "markdown", "content": f"**????**?\n{text}"},
        ],
    }


def _send_due_reminders(force_today: str | None = None, dedup: bool = True) -> dict:
    with _reminder_lock:
        cfg = _db_load_json("feishu_config")
        if not cfg.get("enabled"):
            return {"ok": True, "sent": 0, "skipped": 0, "message": "feishu disabled"}
        app_id = str(cfg.get("app_id", "") or "")
        app_secret = str(cfg.get("app_secret", "") or "")
        open_ids = [str(x).strip() for x in cfg.get("open_ids", []) if str(x).strip()]
        if not app_id or not app_secret or not open_ids:
            return {"ok": False, "sent": 0, "skipped": 0, "message": "feishu config incomplete"}
        shared = _load_shared_state()
        rows = _build_reminder_days(shared)
        today_key = force_today or date.today().isoformat()
        due = [r for r in rows if r.get("remind_date") == today_key]
        if not due:
            return {"ok": True, "sent": 0, "skipped": 0, "message": "no due reminder"}
        log = _db_load_json("reminder_log")
        sent_set = set(log.get("sent_keys", []) if isinstance(log.get("sent_keys"), list) else [])
        token = _feishu_get_token(app_id, app_secret)
        sent = 0
        skipped = 0
        for row in due:
            msg = _build_message(row)
            card = _build_card_content(row)
            for oid in open_ids:
                uniq = f"{today_key}|{oid}|{row.get('promo_start','')}|{row.get('promo_end','')}"
                if dedup and uniq in sent_set:
                    skipped += 1
                    continue
                try:
                    _send_feishu_message(token, oid, "interactive", card)
                except Exception:
                    _send_feishu_message(token, oid, "text", {"text": msg})
                sent_set.add(uniq)
                sent += 1
        if dedup:
            _db_save_json("reminder_log", {"sent_keys": sorted(sent_set)})
        return {"ok": True, "sent": sent, "skipped": skipped, "message": "done"}


def _reminder_worker() -> None:
    while True:
        try:
            _send_due_reminders()
        except Exception:
            pass
        time.sleep(1800)


@app.post("/api/calculate", response_model=CalcResponse)
async def calculate(req: CalcRequest) -> CalcResponse:
    return await asyncio.to_thread(
        _compute,
        req.history_dates,
        req.future_days,
        req.selected_offsets,
        req.marker_overrides,
        req.mode,
        req.preferred_start,
        req.preferred_end,
        req.recommend,
    )


@app.post("/api/amazon/deal", response_model=AmazonDealResponse)
async def parse_amazon_deal(req: AmazonDealRequest) -> AmazonDealResponse:
    url = req.url.strip()
    raw_html = req.html or ""
    if not url and not raw_html.strip():
        raise HTTPException(status_code=400, detail="url or html is required")
    try:
        page_html = raw_html if raw_html.strip() else await asyncio.to_thread(_fetch_amazon_page, url, req.headers)
        parsed = await asyncio.to_thread(_parse_amazon_deal_fields, page_html)
    except urllib.error.HTTPError as e:
        raise HTTPException(status_code=400, detail=f"amazon request failed: {e.code}") from e
    except urllib.error.URLError as e:
        raise HTTPException(status_code=400, detail=f"amazon request failed: {e.reason}") from e
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"amazon parse failed: {e}") from e
    return AmazonDealResponse(url=url, **parsed)


@app.get("/api/asin-monitor/config", response_model=AsinMonitorConfigResponse)
async def get_asin_monitor_config() -> AsinMonitorConfigResponse:
    data = await asyncio.to_thread(_load_asin_monitor_config)
    return AsinMonitorConfigResponse(**data)


@app.post("/api/asin-monitor/config")
async def set_asin_monitor_config(req: AsinMonitorConfigRequest) -> dict:
    current = await asyncio.to_thread(_load_asin_monitor_config)
    current.update(req.model_dump())
    current["asins"] = _normalize_asin_list(current.get("asins", []))
    await asyncio.to_thread(_save_asin_monitor_config, current)
    return {"ok": True}


@app.get("/api/asin-monitor/results", response_model=AsinMonitorResultsResponse)
async def get_asin_monitor_results(limit: int = 200) -> AsinMonitorResultsResponse:
    rows = await asyncio.to_thread(_load_asin_monitor_results)
    rows = rows[: max(1, min(limit, 1000))]
    return AsinMonitorResultsResponse(results=[AsinMonitorResult(**row) for row in rows])


@app.post("/api/asin-monitor/run-now", response_model=AsinMonitorRunResponse)
async def run_asin_monitor_now() -> AsinMonitorRunResponse:
    rows = await asyncio.to_thread(_execute_asin_monitor_run, True)
    return AsinMonitorRunResponse(
        ok=True,
        count=len(rows),
        results=[AsinMonitorResult(**row) for row in rows],
    )


@app.get("/api/asin-monitor/progress")
async def get_asin_monitor_progress() -> dict:
    return dict(_asin_monitor_progress)


@app.get("/api/shared-state", response_model=SharedStateResponse)
async def get_shared_state() -> SharedStateResponse:
    data = await asyncio.to_thread(_load_shared_state)
    return SharedStateResponse(
        records=data.get("records", []),
        product_sim=data.get("product_sim", []),
        future_days=int(data.get("future_days", 90) or 90),
        mode=str(data.get("mode", "even") or "even"),
        custom_start=int(data.get("custom_start", 0) or 0),
        custom_end=int(data.get("custom_end", 14) or 14),
        selected_key=str(data.get("selected_key", "") or ""),
        file_name=str(data.get("file_name", "") or ""),
        marker_mode=str(data.get("marker_mode", "PD") or "PD"),
    )


@app.post("/api/shared-state")
async def set_shared_state(req: SharedStateRequest) -> dict:
    payload = req.model_dump()
    await asyncio.to_thread(_save_shared_state, payload)
    return {"ok": True}


@app.post("/api/upload-source")
async def upload_source(file: UploadFile = File(...)) -> dict:
    content = await file.read()
    await asyncio.to_thread(SOURCE_FILE_PATH.write_bytes, content)
    data = await asyncio.to_thread(_load_shared_state)
    data["file_name"] = file.filename or "source_import.xlsx"
    await asyncio.to_thread(_save_shared_state, data)
    return {"ok": True, "file_name": data["file_name"]}


@app.get("/api/source-file")
async def download_source_file() -> FileResponse:
    if not SOURCE_FILE_PATH.exists():
        raise HTTPException(status_code=404, detail="source file not found")
    data = await asyncio.to_thread(_load_shared_state)
    fname = data.get("file_name") or "source_import.xlsx"
    return FileResponse(
        path=SOURCE_FILE_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=fname,
    )


@app.head("/api/source-file")
async def head_source_file() -> dict:
    if not SOURCE_FILE_PATH.exists():
        raise HTTPException(status_code=404, detail="source file not found")
    return {"ok": True}


@app.get("/api/feishu-config", response_model=FeishuConfigResponse)
async def get_feishu_config() -> FeishuConfigResponse:
    data = await asyncio.to_thread(_db_load_json, "feishu_config")
    return FeishuConfigResponse(
        enabled=bool(data.get("enabled", False)),
        app_id=str(data.get("app_id", "") or ""),
        open_ids=[str(x).strip() for x in data.get("open_ids", []) if str(x).strip()],
        has_secret=bool(data.get("app_secret")),
    )


@app.post("/api/feishu-config")
async def set_feishu_config(req: FeishuConfigRequest) -> dict:
    old = await asyncio.to_thread(_db_load_json, "feishu_config")
    payload = old.copy()
    payload["enabled"] = bool(req.enabled)
    payload["app_id"] = req.app_id.strip()
    payload["open_ids"] = [x.strip() for x in req.open_ids if x.strip()]
    if req.app_secret.strip():
        payload["app_secret"] = req.app_secret.strip()
    await asyncio.to_thread(_db_save_json, "feishu_config", payload)
    return {"ok": True}


@app.get("/api/reminders/preview", response_model=list[ReminderDay])
async def preview_reminders() -> list[ReminderDay]:
    shared = await asyncio.to_thread(_load_shared_state)
    rows = await asyncio.to_thread(_build_reminder_days, shared)
    return [ReminderDay(**x) for x in rows]


@app.post("/api/reminders/send-now")
async def send_reminders_now() -> dict:
    return await asyncio.to_thread(_send_due_reminders, None, False)


@app.get("/api/feishu/departments")
async def feishu_departments() -> dict:
    try:
        cfg = await asyncio.to_thread(_db_load_json, "feishu_config")
        app_id = str(cfg.get("app_id", "") or "")
        app_secret = str(cfg.get("app_secret", "") or "")
        if not app_id or not app_secret:
            raise HTTPException(status_code=400, detail="请先保存飞书 app_id/app_secret")
        token = await asyncio.to_thread(_feishu_get_token, app_id, app_secret)
        root_url = "https://open.feishu.cn/open-apis/contact/v3/departments/0/children?department_id_type=department_id&page_size=50"
        data = await asyncio.to_thread(_feishu_get_json, token, root_url)
        items = data.get("items", []) if isinstance(data, dict) else []
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"加载部门失败: {e}")


@app.get("/api/feishu/users")
async def feishu_users(department_id: str) -> dict:
    try:
        cfg = await asyncio.to_thread(_db_load_json, "feishu_config")
        app_id = str(cfg.get("app_id", "") or "")
        app_secret = str(cfg.get("app_secret", "") or "")
        if not app_id or not app_secret:
            raise HTTPException(status_code=400, detail="请先保存飞书 app_id/app_secret")
        if not department_id.strip():
            raise HTTPException(status_code=400, detail="department_id 不能为空")
        token = await asyncio.to_thread(_feishu_get_token, app_id, app_secret)
        q = urllib.parse.urlencode(
            {
                "department_id": department_id.strip(),
                "department_id_type": "open_department_id",
                "user_id_type": "open_id",
                "page_size": 50,
            }
        )
        url = f"https://open.feishu.cn/open-apis/contact/v3/users/find_by_department?{q}"
        data = await asyncio.to_thread(_feishu_get_json, token, url)
        items = data.get("items", []) if isinstance(data, dict) else []
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"加载成员失败: {e}")


@app.on_event("startup")
async def startup_event() -> None:
    await asyncio.to_thread(_ensure_sqlite_ready)
    th = threading.Thread(target=_reminder_worker, daemon=True, name="feishu-reminder-worker")
    th.start()
    th2 = threading.Thread(target=_asin_monitor_worker, daemon=True, name="amazon-asin-monitor-worker")
    th2.start()


@app.get("/api/feishu/scope-users")
async def feishu_scope_users() -> dict:
    try:
        cfg = await asyncio.to_thread(_db_load_json, "feishu_config")
        app_id = str(cfg.get("app_id", "") or "")
        app_secret = str(cfg.get("app_secret", "") or "")
        if not app_id or not app_secret:
            raise HTTPException(status_code=400, detail="请先保存飞书 app_id/app_secret")
        token = await asyncio.to_thread(_feishu_get_token, app_id, app_secret)
        data = await asyncio.to_thread(_feishu_get_json, token, "https://open.feishu.cn/open-apis/contact/v3/scopes")
        user_ids = data.get("user_ids", []) if isinstance(data, dict) else []
        open_ids = [str(x).strip() for x in user_ids if str(x).strip()]
        items: list[dict] = []
        if open_ids:
            # 先走批量接口，速度更快，信息更完整
            for i in range(0, len(open_ids), 50):
                chunk = open_ids[i : i + 50]
                try:
                    rows = await asyncio.to_thread(_feishu_get_users_batch, token, chunk)
                    got = {x.get("open_id", ""): x for x in rows}
                    for oid in chunk:
                        if oid in got:
                            items.append(got[oid])
                        else:
                            items.append({"open_id": oid, "name": "成员", "avatar": ""})
                except Exception:
                    # 批量失败回退单查，确保不丢人
                    for oid in chunk:
                        try:
                            prof = await asyncio.to_thread(_feishu_get_user_profile, token, oid)
                            items.append(prof)
                        except Exception:
                            items.append({"open_id": oid, "name": "成员", "avatar": ""})
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"加载成员失败: {e}")
